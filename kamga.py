from openpyxl import load_workbook
#load in the workbook
wb = load_workbook("employeedata.xlsx")
ws = wb.active 
for i in range(2,ws.max_row +1): 
    cell = ws.cell(i,1)  #choosing the column i am going to work with
    if "helpinghands.cm" in cell.value:
        updated_email= (cell.value).replace("helpinghands.cm','handsinhands.org") #replacing the old domaine by the new domaine
        ws.cell(i,1).value = updated_email
        #wb.save('new.xlsx')#now we are creating the new file with the ne domaine name  
wb.save("new.xlsx")       
